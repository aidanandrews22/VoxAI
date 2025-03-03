"""
Spreadsheet processor module for extracting text from CSV and Excel spreadsheet files.
"""
import io
import os
import tempfile
from typing import Dict, Any, List

import pandas as pd

from app.core.logging import logger
from app.services.file_processors import FileProcessor


class SpreadsheetProcessor(FileProcessor):
    """
    Base processor for spreadsheet files.
    """
    
    async def process(self, file_content: bytes, file_path: str) -> str:
        """
        Process spreadsheet content and extract text.
        
        Args:
            file_content: Raw spreadsheet file content bytes
            file_path: Path to the spreadsheet file
            
        Returns:
            str: Extracted text content
        """
        raise NotImplementedError("Subclasses must implement this method")
    
    async def get_metadata(self, file_content: bytes, file_path: str) -> Dict[str, Any]:
        """
        Extract metadata from spreadsheet file.
        
        Args:
            file_content: Raw spreadsheet file content bytes
            file_path: Path to the spreadsheet file
            
        Returns:
            Dict[str, Any]: Extracted metadata
        """
        return {
            'file_size': len(file_content),
            'file_extension': os.path.splitext(file_path)[1],
        }
    
    def _format_dataframe_as_text(self, df: pd.DataFrame) -> str:
        """
        Format a pandas DataFrame as readable text.
        
        Args:
            df: DataFrame to format
            
        Returns:
            str: Text representation of the DataFrame
        """
        if df.empty:
            return ""
        
        # Try to infer whether this is tabular data or a list
        if len(df.columns) == 1:
            # Single column - format as a list
            return "\n".join([f"- {item}" for item in df.iloc[:, 0].astype(str).tolist()])
        else:
            # Multiple columns - format as a table with headers
            text_parts = []
            
            # Add header
            header = " | ".join([str(col) for col in df.columns])
            text_parts.append(header)
            text_parts.append("-" * len(header))
            
            # Add rows
            for _, row in df.iterrows():
                text_parts.append(" | ".join([str(val) for val in row.tolist()]))
            
            return "\n".join(text_parts)


class CSVProcessor(SpreadsheetProcessor):
    """
    Processor for CSV files.
    Handles: text/csv
    """
    
    async def process(self, file_content: bytes, file_path: str) -> str:
        """
        Process CSV content and extract text.
        
        Args:
            file_content: Raw CSV file content bytes
            file_path: Path to the CSV file
            
        Returns:
            str: Extracted text content
        """
        try:
            # Load CSV into a pandas DataFrame
            df = pd.read_csv(io.BytesIO(file_content))
            
            # Format the DataFrame as text
            return self._format_dataframe_as_text(df)
        except Exception as e:
            logger.error(f"Error processing CSV file: {e}")
            # Fallback: try to decode as plain text
            try:
                return file_content.decode('utf-8')
            except Exception:
                raise
    
    async def get_metadata(self, file_content: bytes, file_path: str) -> Dict[str, Any]:
        """
        Extract metadata from CSV file.
        
        Args:
            file_content: Raw CSV file content bytes
            file_path: Path to the CSV file
            
        Returns:
            Dict[str, Any]: Extracted metadata
        """
        base_metadata = await super().get_metadata(file_content, file_path)
        
        try:
            # Load CSV into a pandas DataFrame
            df = pd.read_csv(io.BytesIO(file_content))
            
            metadata = {
                'row_count': len(df),
                'column_count': len(df.columns),
                'column_names': df.columns.tolist(),
                'data_types': {col: str(dtype) for col, dtype in df.dtypes.items()},
            }
            
            # Add to base metadata
            base_metadata.update(metadata)
            
            return base_metadata
        except Exception as e:
            logger.error(f"Error extracting CSV metadata: {e}")
            return base_metadata


class ExcelProcessor(SpreadsheetProcessor):
    """
    Processor for Excel files.
    Handles: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    """
    
    async def process(self, file_content: bytes, file_path: str) -> str:
        """
        Process Excel content and extract text.
        
        Args:
            file_content: Raw Excel file content bytes
            file_path: Path to the Excel file
            
        Returns:
            str: Extracted text content
        """
        try:
            # Determine the Excel engine based on file extension
            extension = os.path.splitext(file_path)[1].lower()
            engine = 'openpyxl' if extension == '.xlsx' else 'xlrd'
            
            # Load Excel into pandas DataFrames (one per sheet)
            excel_file = io.BytesIO(file_content)
            
            # Get all sheet names
            sheet_names = pd.ExcelFile(excel_file, engine=engine).sheet_names
            
            excel_file.seek(0)  # Reset file pointer
            
            full_text = []
            
            # Process each sheet
            for sheet_name in sheet_names:
                df = pd.read_excel(excel_file, sheet_name=sheet_name, engine=engine)
                
                if not df.empty:
                    # Add sheet name
                    full_text.append(f"Sheet: {sheet_name}")
                    full_text.append("=" * (len(sheet_name) + 7))  # Underline
                    
                    # Format sheet as text
                    full_text.append(self._format_dataframe_as_text(df))
                    
                    # Add spacing between sheets
                    full_text.append("\n")
            
            return "\n".join(full_text)
        except Exception as e:
            logger.error(f"Error processing Excel file: {e}")
            raise
    
    async def get_metadata(self, file_content: bytes, file_path: str) -> Dict[str, Any]:
        """
        Extract metadata from Excel file.
        
        Args:
            file_content: Raw Excel file content bytes
            file_path: Path to the Excel file
            
        Returns:
            Dict[str, Any]: Extracted metadata
        """
        base_metadata = await super().get_metadata(file_content, file_path)
        
        try:
            # Determine the Excel engine based on file extension
            extension = os.path.splitext(file_path)[1].lower()
            engine = 'openpyxl' if extension == '.xlsx' else 'xlrd'
            
            # Load Excel file
            excel_file = io.BytesIO(file_content)
            xl = pd.ExcelFile(excel_file, engine=engine)
            
            # Extract limited sheet information to prevent metadata size issues
            MAX_SHEETS = 10  # Limit the number of sheets to analyze
            MAX_ROWS_SAMPLE = 100  # Limit row analysis to first N rows
            
            sheet_names = xl.sheet_names[:MAX_SHEETS]
            sheet_info = []
            
            for sheet_name in sheet_names:
                # Only read a sample of rows to avoid large metadata
                df = pd.read_excel(excel_file, sheet_name=sheet_name, engine=engine, nrows=MAX_ROWS_SAMPLE)
                
                # Calculate row count differently to get total count
                if engine == 'openpyxl':
                    try:
                        from openpyxl import load_workbook
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                            temp_file.write(file_content)
                            temp_file.flush()
                            wb = load_workbook(temp_file.name, read_only=True)
                            sheet = wb[sheet_name]
                            row_count = sheet.max_row
                            os.unlink(temp_file.name)
                    except Exception as e:
                        logger.error(f"Error getting exact row count: {e}")
                        row_count = len(df)
                else:
                    row_count = len(df)
                
                sheet_info.append({
                    'name': sheet_name,
                    'row_count': row_count,
                    'column_count': len(df.columns),
                    # Only include column names, not full data types which can be large
                    'column_names': df.columns.tolist()[:30]  # Limit to first 30 columns
                })
                
                excel_file.seek(0)  # Reset file pointer for next sheet
            
            metadata = {
                'sheet_count': len(xl.sheet_names),
                'sheet_names': sheet_names,
                'sheets': sheet_info
            }
            
            # If using openpyxl, try to extract document properties (basic ones only)
            if engine == 'openpyxl':
                # Temporary file required for accessing document properties
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                    temp_file.write(file_content)
                    temp_file.flush()
                    temp_path = temp_file.name
                
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(temp_path)
                    
                    doc_props = {}
                    if wb.properties:
                        props = wb.properties
                        if props.title:
                            doc_props['title'] = props.title
                        if props.creator:
                            doc_props['creator'] = props.creator
                        if props.created:
                            doc_props['created'] = props.created.isoformat()
                        if props.modified:
                            doc_props['modified'] = props.modified.isoformat()
                    
                    if doc_props:
                        metadata['properties'] = doc_props
                    
                except Exception as e:
                    logger.error(f"Error extracting Excel properties: {e}")
                
                os.unlink(temp_path)
            
            # Add to base metadata
            base_metadata.update(metadata)
            
            return base_metadata
        except Exception as e:
            logger.error(f"Error extracting Excel metadata: {e}")
            return base_metadata 